from openpyxl import load_workbook, Workbook
from apikey import upbit
import datetime
import time
import pyupbit

coinname = {'KRW-ETC':'이더리움클래식', 'KRW-XRP':'리플', 'KRW-ETH':'이더리움', 
'KRW-BTC':'비트코인캐시', 'KRW-OMG':'오미세고', 'KRW-EOS':'이오스'}

def write_trade(trade, diff):
    wb = load_workbook('BOT/record.xlsx')
    ws = wb[wb.sheetnames[0]]
    row = []
    day = trade['created_at'].split('T')[0]
    time = trade['created_at'].split('T')[1].split('+')[0]
    row.append(day)
    row.append(time)
    row.append(coinname[trade['market']])
    if trade['side'] == 'ask':
        row.append('매도')
        row.append(diff)
    else :
        row.append('매수')
        row.append(trade['price'])
    ws.append(row)
    wb.save('BOT/record.xlsx')

def write_regularly(coinlist):
    wb = load_workbook('BOT/record.xlsx')
    ws = wb[wb.sheetnames[1]]
    row = []
    day = str(datetime.datetime.now()).split('.')[0]
    d = day.split(' ')[0]
    t = day.split(' ')[1]
    row.append(d)
    row.append(t)
    total = upbit.get_balance("KRW")
    for coin in coinlist:
        price = pyupbit.get_current_price(coin)
        row.append(price)
        total += upbit.get_balance(coin)*price
    row.append(total)
    ws.append(row)
    wb.save('BOT/record.xlsx')
    